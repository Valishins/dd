"""
Модуль предназначен для работы с excel и выгрузки таблицы
"""
from openpyxl import Workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def save_excel(filename: str, rows: list[list[str]]):
    """
    Функция получает название файла и список самолетов с пролетными точками.
    и сохраняет значения в excel файл
    """
    # создание документа
    wb = Workbook()
    # создание вкладки
    wb_active_table = wb.active
    wb_active_table.title = "Общая"
    wb_active_table.append([
        "ID",
        "Месяц",
        "№ пары",
        "Предмет",
        "Группа",
        "Вид занятия",
        "Неделя",
        "Кто",
    ])
    try:
        for row in rows:
            wb_active_table.append(row)
        wb.save(filename)
    except Exception as err:
        print(f"Проблемы с сохранением: { err }")
